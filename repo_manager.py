#!/usr/bin/env python3
"""Manage GitHub repositories from an Excel file.

Two commands:
  export  -  pull your repos into an editable .xlsx (dropdowns for New Visibility /
             Action / Set Description?). Add --descriptions to scan READMEs and
             suggest descriptions; add --check-env to detect and export committed
             root .env files (their contents may include secrets).
  apply   -  read the edited .xlsx and apply changes (preview by default; --yes to execute)

All GitHub operations go through the `gh` CLI, so authentication is whatever
`gh auth login` set up. No GitHub token handling lives in this script.
"""

import argparse
import base64
import concurrent.futures
import datetime
import json
import os
import re
import subprocess
import sys
import zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.datavalidation import DataValidation

# --- configuration ----------------------------------------------------------

JSON_FIELDS = ("name,nameWithOwner,visibility,description,isArchived,isFork,"
               "createdAt,pushedAt,diskUsage,url")
MASS_DELETE_THRESHOLD = 3
DESCRIPTION_MAX = 250  # GitHub allows 350; keep suggestions concise
CLAUDE_MODEL = os.environ.get("REPO_MANAGER_CLAUDE_MODEL", "claude-haiku-4-5")  # cheap Claude (Haiku)
OPENAI_MODEL = os.environ.get("REPO_MANAGER_OPENAI_MODEL", "gpt-4o-mini")  # cheap OpenAI option

HEADERS = [
    "Owner", "Repo", "URL", "Current Visibility", "New Visibility", "Action",
    "Has README", "Current Description", "New Description", "Set Description?",
    "Archived", "Fork", "Created", "Last Push", "Size (KB)",
    "Has .env", ".env Content", "New .env Content", ".env Action",
]
# Cells are read-only (locked) unless their header is in EDITABLE.
EDITABLE = {"New Visibility", "Action", "New Description", "Set Description?",
            "New .env Content", ".env Action"}
COL = {h: i + 1 for i, h in enumerate(HEADERS)}  # 1-based column index by header

COL_WIDTHS = {
    "Owner": 16, "Repo": 28, "URL": 50, "Current Visibility": 16,
    "New Visibility": 15, "Action": 9, "Has README": 12,
    "Current Description": 40, "New Description": 45, "Set Description?": 15,
    "Archived": 9, "Fork": 7, "Created": 12, "Last Push": 12, "Size (KB)": 9,
    "Has .env": 10, ".env Content": 45, "New .env Content": 45, ".env Action": 12,
}

SNAPSHOT_SHEET = "_snapshot"
DATA_SHEET = "repos"


# --- gh helpers -------------------------------------------------------------

class GhError(Exception):
    def __init__(self, cmd, code, stdout, stderr):
        self.cmd = cmd
        self.code = code
        self.stdout = stdout
        self.stderr = stderr
        super().__init__(f"gh exited {code}: {' '.join(cmd)}\n{(stderr or '').strip()}")


def run_gh(args, check=True, timeout=120, input_data=None):
    """Run `gh <args>` and return the CompletedProcess. `input_data` is sent on stdin."""
    cmd = ["gh"] + args
    try:
        result = subprocess.run(cmd, capture_output=True, text=True,
                                encoding="utf-8", errors="replace", timeout=timeout,
                                input=input_data)
    except FileNotFoundError:
        sys.exit("ERROR: GitHub CLI 'gh' not found on PATH.\n"
                 "Install it from https://cli.github.com/ and run 'gh auth login'.")
    except subprocess.TimeoutExpired:
        raise GhError(cmd, -1, "", f"command timed out after {timeout}s")
    if check and result.returncode != 0:
        raise GhError(cmd, result.returncode, result.stdout, result.stderr)
    return result


def ensure_auth():
    r = run_gh(["auth", "status"], check=False)
    if r.returncode != 0:
        sys.exit("ERROR: not authenticated with GitHub.\nRun 'gh auth login' first.\n" + (r.stderr or ""))


def get_login():
    return run_gh(["api", "user", "-q", ".login"]).stdout.strip()


def list_repos(owner):
    """Return a list of repo dicts for `owner`, visibility normalised to lowercase."""
    r = run_gh(["repo", "list", owner, "--limit", "4000", "--json", JSON_FIELDS])
    try:
        data = json.loads(r.stdout or "[]")
    except json.JSONDecodeError:
        sys.exit("ERROR: could not parse the repo list returned by gh.\n"
                 "Try 'gh repo list --limit 5' to check your gh setup.")
    for repo in data:
        repo["visibility"] = (repo.get("visibility") or "").lower()
    return data


# --- README scan + heuristic summary ----------------------------------------

_BADGE_RE = re.compile(r"^\s*\[?!\[")            # image / linked image (badges) at line start
_IMG_RE = re.compile(r"!\[[^\]]*\]\([^)]*\)")    # ![alt](url)
_LINK_RE = re.compile(r"\[([^\]]+)\]\([^)]*\)")  # [text](url) -> text
_HTML_TAG_RE = re.compile(r"<[^>]+>")
_EMPHASIS_RE = re.compile(r"[*_`~]+")
_URL_RE = re.compile(r"https?://\S+")            # bare URLs
_HR_RE = re.compile(r"^[-*_]{3,}$")              # horizontal rule
_HTML_LINE_RE = re.compile(r"^<[^>]+>$")         # a line that is only an HTML tag


def _fetch_raw(endpoint, timeout=30):
    """GET a GitHub endpoint as raw text, or None if it 404s / fails / times out.

    Bytes are decoded UTF-8 -> Windows-1252 -> Latin-1 so files in legacy European
    encodings keep their accents (Latin-1 never raises, so decoding always succeeds).
    """
    cmd = ["gh", "api", endpoint, "-H", "Accept: application/vnd.github.raw"]
    try:
        result = subprocess.run(cmd, capture_output=True, timeout=timeout)
    except (FileNotFoundError, OSError, subprocess.TimeoutExpired):
        return None
    if result.returncode != 0:
        return None
    for enc in ("utf-8", "cp1252"):
        try:
            return result.stdout.decode(enc)
        except UnicodeDecodeError:
            continue
    return result.stdout.decode("latin-1")


def fetch_readme(name_with_owner):
    """Return the README text for a repo, or None if it has no README."""
    return _fetch_raw(f"repos/{name_with_owner}/readme")


def fetch_env(name_with_owner):
    """Return the repo-root .env text, or None if absent. May contain secrets."""
    return _fetch_raw(f"repos/{name_with_owner}/contents/.env")


def env_sha(name_with_owner):
    """Return the blob sha of the repo-root .env, or None if it doesn't exist."""
    r = run_gh(["api", f"repos/{name_with_owner}/contents/.env", "--jq", ".sha"], check=False)
    sha = r.stdout.strip()
    return sha if r.returncode == 0 and sha else None


def put_env(name_with_owner, content, message):
    """Create or update the repo-root .env (commits to the default branch).

    The request body is sent on stdin so large content can't blow the command-line
    length limit.
    """
    body = {"message": message,
            "content": base64.b64encode(content.encode("utf-8")).decode("ascii")}
    sha = env_sha(name_with_owner)
    if sha:
        body["sha"] = sha
    run_gh(["api", "--method", "PUT", f"repos/{name_with_owner}/contents/.env", "--input", "-"],
           input_data=json.dumps(body))


def delete_env(name_with_owner, message):
    """Delete the repo-root .env (commits to the default branch). Raises if absent."""
    sha = env_sha(name_with_owner)
    if not sha:
        raise GhError(["gh", "api"], -1, "", "no .env found to delete")
    body = {"message": message, "sha": sha}
    run_gh(["api", "--method", "DELETE", f"repos/{name_with_owner}/contents/.env", "--input", "-"],
           input_data=json.dumps(body))


def _clean_md(text):
    """Strip markdown/HTML noise and bare URLs from a line of prose."""
    text = _IMG_RE.sub("", text)
    text = _LINK_RE.sub(r"\1", text)          # [label](url) -> label
    text = _URL_RE.sub("", text)              # drop bare URLs
    text = _HTML_TAG_RE.sub("", text)
    text = _EMPHASIS_RE.sub("", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text.strip("-*•> \t").strip()


def _first_sentence(text):
    """Return the first sentence (up to . ! or ?), else the whole string."""
    m = re.search(r"[.!?](?:\s|$)", text)
    return text[:m.start() + 1].strip() if m else text


def summarize_readme(text, max_len=DESCRIPTION_MAX):
    """Heuristically pull a concise one-line description from README markdown.

    Skips badges, headings, code, blockquotes, HTML-only lines and horizontal rules,
    then takes the first substantial prose paragraph and prefers its first sentence.
    """
    if not text:
        return ""
    paragraphs, current, in_code = [], [], False
    for raw in text.splitlines():
        line = raw.strip()
        if line.startswith("```") or line.startswith("~~~"):
            in_code = not in_code
            continue
        if in_code:
            continue
        if not line:
            if current:
                paragraphs.append(" ".join(current))
                current = []
            continue
        if (line.startswith("#") or line.startswith(">") or line.startswith("<!--")
                or _BADGE_RE.match(line) or _HR_RE.match(line) or _HTML_LINE_RE.match(line)):
            if current:
                paragraphs.append(" ".join(current))
                current = []
            continue
        line = re.sub(r"^([-*+]|\d+[.)])\s+", "", line)  # strip leading list marker
        current.append(line)
    if current:
        paragraphs.append(" ".join(current))

    for para in paragraphs:
        cleaned = _clean_md(para)
        if len(cleaned) < 20:
            continue
        sentence = _first_sentence(cleaned)
        candidate = sentence if len(sentence) >= 40 else cleaned
        if len(candidate) > max_len:
            candidate = candidate[:max_len].rsplit(" ", 1)[0].rstrip(",.;:") + "..."
        return candidate
    return ""


_AI_SYSTEM_PROMPT = (
    "You write the GitHub 'About' description for a repository, given its README. "
    "Reply with ONE plain-text line of at most 160 characters describing what the "
    "project is or does: no markdown, no surrounding quotes. If the README does not "
    "contain enough meaningful information to describe the project, reply with "
    "exactly: NONE"
)


def _make_claude_call():
    """Return call(text)->str using Claude (cheap model). Validates SDK + key."""
    try:
        import anthropic
    except ImportError:
        sys.exit("ERROR: --ai (claude) needs the 'anthropic' package. Install it with:\n"
                 "  pip install anthropic")
    if not os.environ.get("ANTHROPIC_API_KEY"):
        sys.exit("ERROR: --ai (claude) needs ANTHROPIC_API_KEY.\n"
                 "Set it in your environment or in a .env file.")
    client = anthropic.Anthropic()

    def call(text):
        resp = client.messages.create(
            model=CLAUDE_MODEL, max_tokens=200, system=_AI_SYSTEM_PROMPT,
            messages=[{"role": "user", "content": text}],
        )
        return "".join(b.text for b in resp.content if b.type == "text")
    return call


def _make_openai_call():
    """Return call(text)->str using OpenAI (cheap model). Validates SDK + key."""
    try:
        import openai
    except ImportError:
        sys.exit("ERROR: --ai-provider openai needs the 'openai' package. Install it with:\n"
                 "  pip install openai")
    if not os.environ.get("OPENAI_API_KEY"):
        sys.exit("ERROR: --ai-provider openai needs OPENAI_API_KEY.\n"
                 "Set it in your environment or in a .env file.")
    client = openai.OpenAI()

    def call(text):
        resp = client.chat.completions.create(
            model=OPENAI_MODEL, max_tokens=200,
            messages=[{"role": "system", "content": _AI_SYSTEM_PROMPT},
                      {"role": "user", "content": text}],
        )
        return resp.choices[0].message.content or ""
    return call


def make_ai_summarizer(provider="claude"):
    """Return summarize(text)->str backed by the chosen AI provider.

    Uses the heuristic as a cheap "is this worth summarizing?" pre-check, and
    falls back to the heuristic on any API error.
    """
    call = _make_openai_call() if provider == "openai" else _make_claude_call()

    def summarize(text):
        if not summarize_readme(text):       # no real prose -> not worth an API call
            return ""
        try:
            out = call(text[:6000])
        except Exception:
            return summarize_readme(text)    # graceful fallback to the heuristic
        out = " ".join(out.split()).strip().strip('"').strip()
        if not out or out.upper() == "NONE":
            return ""
        return out[:DESCRIPTION_MAX]

    return summarize


ENV_CONTENT_MAX = 30000  # Excel cell limit is 32767 chars; stay safely under


def scan_repos(repos, want_readme, want_env, summarize_fn=summarize_readme, max_workers=8):
    """Return {nameWithOwner: {has_readme, suggestion, has_env, env_content}}.

    Only the requested checks run; everything else stays at its default. Each
    repo's checks are isolated so one failure can't abort the whole scan.
    """
    def work(repo):
        nwo = repo["nameWithOwner"]
        info = {"has_readme": False, "suggestion": "", "has_env": False, "env_content": ""}
        if want_readme:
            try:
                text = fetch_readme(nwo)
            except Exception:
                text = None
            if text is not None:
                info["has_readme"] = True
                # Only summarize repos lacking a description (saves AI calls).
                if not (repo.get("description") or "").strip():
                    try:
                        info["suggestion"] = summarize_fn(text)
                    except Exception:
                        info["suggestion"] = ""  # never let one repo abort the scan
        if want_env:
            try:
                content = fetch_env(nwo)
            except Exception:
                content = None
            if content is not None:
                info["has_env"] = True
                if len(content) > ENV_CONTENT_MAX:
                    content = content[:ENV_CONTENT_MAX] + "\n...[truncated]"
                info["env_content"] = content
        return nwo, info

    result = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as ex:
        for nwo, info in ex.map(work, repos):
            result[nwo] = info
    return result


# --- export -----------------------------------------------------------------

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
LOCKED_FILL = PatternFill("solid", fgColor="F2F2F2")
PUBLIC_FILL = PatternFill("solid", fgColor="E2EFDA")
PRIVATE_FILL = PatternFill("solid", fgColor="FCE4D6")


def _add_list_validation(ws, header, last_row, items, error, prompt):
    dv = DataValidation(type="list", formula1=f'"{items}"', allow_blank=False)
    dv.error = error
    dv.prompt = prompt
    ws.add_data_validation(dv)
    letter = get_column_letter(COL[header])
    dv.add(f"{letter}2:{letter}{last_row}")


def _write_cell(ws, row, col, value):
    """Write a cell, storing strings that begin with '=' as literal text, not a formula."""
    c = ws.cell(row=row, column=col, value=value)
    if isinstance(value, str) and value.startswith("="):
        c.data_type = "s"
    return c


def cmd_export(args):
    ensure_auth()
    owner = args.owner or get_login()
    print(f"Fetching repos for '{owner}' via gh ...")
    repos = list_repos(owner)
    # Newest first by creation date; fall back to name when dates are equal/missing.
    repos.sort(key=lambda r: r["nameWithOwner"].lower())
    repos.sort(key=lambda r: r.get("createdAt") or "", reverse=True)
    print(f"  {len(repos)} repos found.")

    scan = {}
    ai_enabled = args.ai or args.ai_provider is not None
    provider = args.ai_provider or "claude"
    want_readme = args.descriptions or ai_enabled
    want_env = args.check_env
    summarize_fn = make_ai_summarizer(provider) if ai_enabled else summarize_readme
    if want_readme or want_env:
        targets = (["READMEs"] if want_readme else []) + ([".env files"] if want_env else [])
        if ai_enabled:
            model = OPENAI_MODEL if provider == "openai" else CLAUDE_MODEL
            targets[0] = f"READMEs (AI via {provider}: {model})"
        print(f"Scanning {' and '.join(targets)} (one request per repo) ...")
        scan = scan_repos(repos, want_readme, want_env, summarize_fn=summarize_fn)
        if want_readme:
            have = sum(1 for v in scan.values() if v["has_readme"])
            print(f"  {have}/{len(repos)} repos have a README.")
        if want_env:
            have_env = sum(1 for v in scan.values() if v["has_env"])
            print(f"  {have_env}/{len(repos)} repos have a committed .env in the root.")
            if have_env:
                print("  WARNING: .env contents were written to the spreadsheet and may contain\n"
                      "           secrets. Keep the file private, do not share or commit it, and\n"
                      "           rotate any exposed credentials.")

    wb = Workbook()
    ws = wb.active
    ws.title = DATA_SHEET

    for head in HEADERS:
        c = ws.cell(row=1, column=COL[head], value=head)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(vertical="center")

    for i, repo in enumerate(repos):
        row = i + 2
        vis = repo["visibility"]
        vis_fill = PUBLIC_FILL if vis == "public" else PRIVATE_FILL
        nwo = repo["nameWithOwner"]
        current_desc = repo.get("description") or ""

        info = scan.get(nwo, {})
        if want_readme:
            has_readme = info.get("has_readme", False)
            suggestion = info.get("suggestion", "")
            has_readme_val = "TRUE" if has_readme else "FALSE"
        else:
            has_readme, suggestion, has_readme_val = False, "", "not scanned"

        if want_env:
            has_env_val = "TRUE" if info.get("has_env") else "FALSE"
            env_content = info.get("env_content", "")
        else:
            has_env_val, env_content = "not scanned", ""
        new_env_content = env_content if want_env else ""

        if want_readme and has_readme and suggestion and not current_desc:
            new_desc = suggestion
        else:
            new_desc = current_desc

        values = {
            "Owner": nwo.split("/")[0],
            "Repo": repo["name"],
            "URL": repo["url"],
            "Current Visibility": vis,
            "New Visibility": vis,
            "Action": "keep",
            "Has README": has_readme_val,
            "Has .env": has_env_val,
            "Current Description": current_desc,
            "New Description": new_desc,
            "Set Description?": "keep",
            "Archived": "TRUE" if repo.get("isArchived") else "FALSE",
            "Fork": "TRUE" if repo.get("isFork") else "FALSE",
            "Created": (repo.get("createdAt") or "")[:10],
            "Last Push": (repo.get("pushedAt") or "")[:10],
            "Size (KB)": repo.get("diskUsage") or 0,
            ".env Content": env_content,
            "New .env Content": new_env_content,
            ".env Action": "keep",
        }
        for head in HEADERS:
            c = _write_cell(ws, row, COL[head], values[head])
            if head not in EDITABLE:
                c.fill = LOCKED_FILL  # shade reference columns (not protected, just a hint)
        ws.cell(row=row, column=COL["Current Visibility"]).fill = vis_fill
        ws.cell(row=row, column=COL["New Visibility"]).fill = vis_fill
        for env_col in (".env Content", "New .env Content"):
            ws.cell(row=row, column=COL[env_col]).alignment = Alignment(
                wrap_text=True, vertical="top")

    last_row = max(len(repos) + 1, 2)

    _add_list_validation(ws, "New Visibility", last_row, "public,private",
                         "Choose 'public' or 'private'.", "Target visibility for this repo.")
    _add_list_validation(ws, "Action", last_row, "keep,delete",
                         "Choose 'keep' or 'delete'.", "Set to 'delete' to remove the repo on apply.")
    _add_list_validation(ws, "Set Description?", last_row, "keep,set",
                         "Choose 'keep' or 'set'.", "Set to 'set' to update the description on apply.")
    _add_list_validation(ws, ".env Action", last_row, "keep,update,delete",
                         "Choose 'keep', 'update', or 'delete'.",
                         "'update' commits New .env Content; 'delete' removes the .env file.")

    for head in HEADERS:
        ws.column_dimensions[get_column_letter(COL[head])].width = COL_WIDTHS[head]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{last_row}"

    snap = wb.create_sheet(SNAPSHOT_SHEET)
    snap.append(["nameWithOwner", "visibility", "description"])
    for s_row, repo in enumerate(repos, start=2):
        _write_cell(snap, s_row, 1, repo["nameWithOwner"])
        _write_cell(snap, s_row, 2, repo["visibility"])
        _write_cell(snap, s_row, 3, repo.get("description") or "")
    snap.sheet_state = "hidden"

    try:
        wb.save(args.file)
    except PermissionError:
        sys.exit(f"ERROR: could not write {args.file}. "
                 f"If it's open in Excel, close it and run export again.")
    print(f"Wrote {args.file}")
    print("Next: edit the editable (white) columns, then run:")
    print(f"  python repo_manager.py apply {args.file}            # preview")
    print(f"  python repo_manager.py apply {args.file} --yes      # execute")
    tips = []
    if not args.descriptions and not ai_enabled:
        tips.append("--descriptions or --ai (suggest descriptions from READMEs)")
    if not args.check_env:
        tips.append("--check-env (detect + export committed .env files)")
    if tips:
        print("Tip: add " + " or ".join(tips) + ".")


# --- apply ------------------------------------------------------------------

def _cell(value):
    return value if value is not None else ""


def read_rows(path):
    if not os.path.exists(path):
        sys.exit(f"ERROR: file not found: {path}\nRun 'export' first.")
    try:
        wb = load_workbook(path)
    except (InvalidFileException, zipfile.BadZipFile, PermissionError) as e:
        sys.exit(f"ERROR: could not open {path} ({e}).\n"
                 f"Make sure it's an .xlsx created by 'export' and not open in Excel.")
    if DATA_SHEET not in wb.sheetnames:
        sys.exit(f"ERROR: sheet '{DATA_SHEET}' not found in {path}. Was this made by 'export'?")
    ws = wb[DATA_SHEET]

    headers = [(_cell(c.value)).strip() if isinstance(c.value, str) else _cell(c.value) for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    for req in ("Owner", "Repo", "Current Visibility", "New Visibility", "Action"):
        if req not in idx:
            sys.exit(f"ERROR: column '{req}' missing in {path}. Re-export the file.")

    def get(row, name):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        owner = get(r, "Owner")
        repo = get(r, "Repo")
        if not owner and not repo:
            continue
        rows.append({
            "owner": str(owner).strip(),
            "repo": str(repo).strip(),
            "nameWithOwner": f"{str(owner).strip()}/{str(repo).strip()}",
            "current_visibility": str(_cell(get(r, "Current Visibility"))).strip().lower(),
            "new_visibility": str(_cell(get(r, "New Visibility"))).strip().lower(),
            "action": (str(_cell(get(r, "Action"))).strip().lower() or "keep"),
            "archived": str(_cell(get(r, "Archived"))).strip().upper() == "TRUE",
            "current_description": str(_cell(get(r, "Current Description"))).strip(),
            "new_description": str(_cell(get(r, "New Description"))).strip(),
            "set_description": (str(_cell(get(r, "Set Description?"))).strip().lower() or "keep"),
            "current_env_content": str(_cell(get(r, ".env Content"))),
            "new_env_content": str(_cell(get(r, "New .env Content"))),
            "env_action": (str(_cell(get(r, ".env Action"))).strip().lower() or "keep"),
        })

    snapshot = {}
    if SNAPSHOT_SHEET in wb.sheetnames:
        for s in wb[SNAPSHOT_SHEET].iter_rows(min_row=2, values_only=True):
            if s and s[0]:
                snapshot[s[0]] = {
                    "visibility": str(_cell(s[1])).strip().lower() if len(s) > 1 else "",
                    "description": str(_cell(s[2])).strip() if len(s) > 2 else "",
                }
    return rows, snapshot


def _add_once(lst, item):
    if item not in lst:
        lst.append(item)


# Change types that actually touch GitHub, in execution order.
ACTIONABLE_KEYS = ("visibility", "descriptions", "env_updates", "env_deletes", "deletes")


def compute_changes(rows):
    """Group requested actions into a dict of row-lists keyed by change type."""
    changes = {k: [] for k in ACTIONABLE_KEYS}
    changes["skipped_archived"] = []
    for row in rows:
        if row["action"] == "delete":
            changes["deletes"].append(row)
            continue
        archived = row["archived"]

        nv = row["new_visibility"]
        if nv and nv != row["current_visibility"] and nv in ("public", "private"):
            if archived:
                _add_once(changes["skipped_archived"], row)
            else:
                changes["visibility"].append(row)

        if row["set_description"] == "set":
            nd = row["new_description"]
            if nd and nd != row["current_description"]:
                if archived:
                    _add_once(changes["skipped_archived"], row)
                else:
                    changes["descriptions"].append(row)

        env_action = row["env_action"]
        if env_action == "update":
            new_env = row["new_env_content"]
            if new_env.strip() and new_env != row["current_env_content"]:
                if archived:
                    _add_once(changes["skipped_archived"], row)
                else:
                    changes["env_updates"].append(row)
        elif env_action == "delete":
            if archived:
                _add_once(changes["skipped_archived"], row)
            else:
                changes["env_deletes"].append(row)
    return changes


def fetch_live_map(owners):
    live = {}
    for owner in owners:
        for repo in list_repos(owner):
            live[repo["nameWithOwner"]] = {
                "visibility": repo["visibility"],
                "description": repo.get("description") or "",
            }
    return live


def print_summary(changes, warnings):
    print("\n=== Planned changes ===")
    vis = changes["visibility"]
    to_priv = [r for r in vis if r["new_visibility"] == "private"]
    to_pub = [r for r in vis if r["new_visibility"] == "public"]
    print(f"Visibility: {len(to_priv)} public->private, {len(to_pub)} private->public")
    for r in vis:
        print(f"  ~ {r['nameWithOwner']}: {r['current_visibility']} -> {r['new_visibility']}")

    print(f"Descriptions to set: {len(changes['descriptions'])}")
    for r in changes["descriptions"]:
        preview = r["new_description"]
        if len(preview) > 60:
            preview = preview[:60] + "..."
        print(f"  = {r['nameWithOwner']}: \"{preview}\"")

    # Never print .env contents (they are secrets) - only sizes.
    n_env_updates = len(changes["env_updates"])
    note = "  (commits .env contents into git)" if n_env_updates else ""
    print(f".env updates: {n_env_updates}{note}")
    for r in changes["env_updates"]:
        print(f"  e {r['nameWithOwner']}  ({len(r['new_env_content'])} chars)")
    print(f".env deletions: {len(changes['env_deletes'])}")
    for r in changes["env_deletes"]:
        print(f"  x {r['nameWithOwner']}")

    print(f"Delete repo: {len(changes['deletes'])}")
    for r in changes["deletes"]:
        print(f"  - {r['nameWithOwner']}")

    if changes["skipped_archived"]:
        print(f"Skipped (archived, can't edit): {len(changes['skipped_archived'])}")
        for r in changes["skipped_archived"]:
            print(f"  . {r['nameWithOwner']}")
    if warnings:
        print("Warnings:")
        for w in warnings:
            print(f"  ! {w}")


def write_log(lines, ok, fail):
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    path = f"repo_manager_{ts}.log"
    with open(path, "w", encoding="utf-8") as f:
        f.write(f"repo_manager apply run at {ts}\n\n")
        f.write("\n".join(lines))
        f.write(f"\n\nSummary: {ok} succeeded, {fail} failed.\n")
    print(f"Log written to {path}")


def cmd_apply(args):
    ensure_auth()
    rows, _snapshot = read_rows(args.file)
    changes = compute_changes(rows)

    owners = sorted({r["owner"] for r in rows if r["owner"]})
    print(f"Checking live state on GitHub for: {', '.join(owners) or '(none)'} ...")
    live = fetch_live_map(owners)

    warnings = []
    for r in rows:
        nwo = r["nameWithOwner"]
        nv = r["new_visibility"]
        if nv and nv != r["current_visibility"] and nv not in ("public", "private"):
            warnings.append(f"{nwo}: New Visibility '{nv}' is not 'public'/'private' - ignored")
        if r["action"] not in ("keep", "delete"):
            warnings.append(f"{nwo}: Action '{r['action']}' unrecognized - treated as 'keep'")
        if r["set_description"] not in ("keep", "set"):
            warnings.append(f"{nwo}: Set Description? '{r['set_description']}' unrecognized - treated as 'keep'")
        if r["env_action"] not in ("keep", "update", "delete"):
            warnings.append(f"{nwo}: .env Action '{r['env_action']}' unrecognized - treated as 'keep'")
        if r["set_description"] == "set" and not r["new_description"]:
            warnings.append(f"{nwo}: 'Set Description?' is 'set' but New Description is empty - skipped")
        if r["env_action"] == "update" and not r["new_env_content"].strip():
            warnings.append(f"{nwo}: '.env Action' is 'update' but New .env Content is empty - skipped")

    actionable = [r for k in ACTIONABLE_KEYS for r in changes[k]]
    for r in actionable:
        nwo = r["nameWithOwner"]
        if nwo not in live:
            warnings.append(f"{nwo}: not found on GitHub (already deleted/renamed) - will skip")
            r["_missing"] = True
        elif r["current_visibility"] and live[nwo]["visibility"] != r["current_visibility"]:
            warnings.append(f"{nwo}: file says '{r['current_visibility']}' but GitHub says "
                            f"'{live[nwo]['visibility']}' (spreadsheet may be stale)")
    for k in ACTIONABLE_KEYS:
        changes[k] = [r for r in changes[k] if not r.get("_missing")]

    warnings = list(dict.fromkeys(warnings))  # de-dup: a row can be in several lists
    print_summary(changes, warnings)

    if not any(changes[k] for k in ACTIONABLE_KEYS):
        print("\nNothing to do.")
        return

    if not args.yes:
        print("\nDRY RUN - no changes made. Re-run with --yes to apply.")
        return

    if len(changes["deletes"]) > MASS_DELETE_THRESHOLD and not args.allow_mass_delete:
        sys.exit(f"\nERROR: {len(changes['deletes'])} deletions requested "
                 f"(> {MASS_DELETE_THRESHOLD}).\n"
                 f"Re-run with --allow-mass-delete if you really mean it.")

    if changes["env_updates"]:
        print("\nNOTE: .env updates commit file contents (often secrets) into the repos' "
              "git history.")

    if changes["deletes"] and not args.force:
        print("\nThe following repos will be PERMANENTLY DELETED:")
        for r in changes["deletes"]:
            print(f"  - {r['nameWithOwner']}")
        try:
            answer = input("Type DELETE to confirm (anything else skips deletions): ").strip()
        except EOFError:
            answer = ""
        if answer != "DELETE":
            print("Deletions cancelled. Other changes (if any) will still proceed.")
            changes["deletes"] = []

    log_lines, ok, fail = [], 0, 0

    def record(line, success):
        nonlocal ok, fail
        if success:
            ok += 1
        else:
            fail += 1
        print(line)
        log_lines.append(line)

    for r in changes["visibility"]:
        nwo, target = r["nameWithOwner"], r["new_visibility"]
        try:
            run_gh(["repo", "edit", nwo, "--visibility", target,
                    "--accept-visibility-change-consequences"])
            record(f"OK    visibility {r['current_visibility']}->{target}  {nwo}", True)
        except GhError as e:
            record(f"FAIL  visibility {nwo}: {(e.stderr or '').strip()}", False)

    for r in changes["descriptions"]:
        nwo = r["nameWithOwner"]
        try:
            run_gh(["repo", "edit", nwo, "--description", r["new_description"]])
            record(f"OK    description set  {nwo}", True)
        except GhError as e:
            record(f"FAIL  description {nwo}: {(e.stderr or '').strip()}", False)

    for r in changes["env_updates"]:
        nwo = r["nameWithOwner"]
        try:
            put_env(nwo, r["new_env_content"], "Update .env (via repo_manager)")
            record(f"OK    .env updated  {nwo}", True)
        except GhError as e:
            record(f"FAIL  .env update {nwo}: {(e.stderr or '').strip()}", False)

    for r in changes["env_deletes"]:
        nwo = r["nameWithOwner"]
        try:
            delete_env(nwo, "Remove .env (via repo_manager)")
            record(f"OK    .env deleted  {nwo}", True)
        except GhError as e:
            record(f"FAIL  .env delete {nwo}: {(e.stderr or '').strip()}", False)

    for r in changes["deletes"]:
        nwo = r["nameWithOwner"]
        try:
            run_gh(["repo", "delete", nwo, "--yes"])
            record(f"OK    repo deleted  {nwo}", True)
        except GhError as e:
            record(f"FAIL  repo delete {nwo}: {(e.stderr or '').strip()}", False)

    if log_lines:
        write_log(log_lines, ok, fail)
    print(f"\nDone: {ok} succeeded, {fail} failed.")
    print(f"Tip: re-run 'python repo_manager.py export {args.file}' to refresh the spreadsheet.")
    if fail:
        sys.exit(1)


# --- cli --------------------------------------------------------------------

def load_dotenv(path=".env"):
    """Load KEY=VALUE pairs from a local .env into os.environ (without overriding).

    Lets you keep ANTHROPIC_API_KEY / OPENAI_API_KEY in a gitignored .env file.
    """
    if not os.path.exists(path):
        return
    try:
        with open(path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                if line.startswith("export "):
                    line = line[len("export "):]
                key, _, val = line.partition("=")
                key = key.strip()
                if key and key not in os.environ:
                    os.environ[key] = val.strip().strip('"').strip("'")
    except OSError:
        pass


def main():
    # Never crash just because a description has an emoji/CJK char the console
    # can't encode; degrade to a replacement char instead.
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8", errors="replace")
        except (AttributeError, ValueError):
            pass

    load_dotenv()  # pick up ANTHROPIC_API_KEY / OPENAI_API_KEY from a local .env

    parser = argparse.ArgumentParser(
        description="Manage GitHub repos from an Excel file (export / apply).")
    sub = parser.add_subparsers(dest="command", required=True)

    pe = sub.add_parser("export", help="Export your repos to an editable Excel file.")
    pe.add_argument("file", nargs="?", default="repos.xlsx", help="output .xlsx (default: repos.xlsx)")
    pe.add_argument("--owner", help="GitHub owner (default: your authenticated login).")
    pe.add_argument("--descriptions", action="store_true",
                    help="Scan READMEs to fill 'Has README' and suggest descriptions "
                         "(built-in heuristic) for repos that don't have one.")
    pe.add_argument("--ai", action="store_true",
                    help="Use an LLM instead of the heuristic to write descriptions "
                         "(implies a README scan).")
    pe.add_argument("--ai-provider", choices=["claude", "openai"], default=None,
                    help="LLM provider for --ai (default: claude). Keys come from the "
                         "environment or a .env file (ANTHROPIC_API_KEY / OPENAI_API_KEY).")
    pe.add_argument("--check-env", action="store_true",
                    help="Detect a committed root .env file per repo and export its "
                         "contents (WARNING: may include secrets) into the spreadsheet.")
    pe.set_defaults(func=cmd_export)

    pa = sub.add_parser("apply", help="Apply changes from the Excel file (preview unless --yes).")
    pa.add_argument("file", nargs="?", default="repos.xlsx", help="input .xlsx (default: repos.xlsx)")
    pa.add_argument("--yes", action="store_true", help="Actually perform changes (default: preview).")
    pa.add_argument("--allow-mass-delete", action="store_true",
                    help=f"Permit more than {MASS_DELETE_THRESHOLD} deletions in one run.")
    pa.add_argument("--force", action="store_true", help="Skip the interactive DELETE confirmation.")
    pa.set_defaults(func=cmd_apply)

    args = parser.parse_args()
    try:
        args.func(args)
    except GhError as e:
        sys.exit(f"ERROR: {e}")
    except KeyboardInterrupt:
        sys.exit("\nAborted.")


if __name__ == "__main__":
    main()
